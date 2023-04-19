import os
import pandas as pd #requires 'pip install pandas fsspec'
import re
import openpyxl #requires 'pip install openpyxl'
import extract_msg #requires 'pip install extract_msg'
from extract_msg.exceptions import InvalidFileFormatError

# Function to read the parameters from the txt file
def read_params_from_txt_file(file_path):
    params = {}
    with open(file_path, 'r') as f:
        for line in f:
            if line.strip():
                key, value = line.strip().split(" = ", 1)
                params[key.strip()] = value.strip()
    return params

# Main execution

# Load the parameters from the text file
params_file_path = r"C:\Mis Datos en Local\temporal\python\email-automation-classify-params.txt"
params = read_params_from_txt_file(params_file_path)

# Set the directory you want to search in
dir_path = params['dir_path']

# Load the existing Excel file as a DataFrame
excel_path = params['excel_path']
try:
    df_existing = pd.read_excel(excel_path)
except FileNotFoundError:
    df_existing = pd.DataFrame(columns=["Subject", "Path", "Sender", "Recipients", "Archive", "Date"])

# Get the column widths from the existing Excel file
if os.path.exists(excel_path):
    wb_existing = openpyxl.load_workbook(excel_path)
    ws_existing = wb_existing.active
    column_widths = [ws_existing.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width for i in range(len(df_existing.columns))]

# Create a list to store the file names and paths
file_list = []

# Initialize a counter for processed emails
processed_emails = 0

# Iterate over all the files in the directory
for subdir, dirs, files in os.walk(dir_path):
    for file in files:

        # Check if the file has a .msg extension
        if file.endswith(".msg"):

            # Get the full path to the file
            file_path = os.path.join(subdir, file)

            # Check if the file path is too long (for example, more than 260 characters)
            if len(file_path) > 260:
                print(f"Skipping file with long path: {file_path}")
                continue

            # Check if the file is already in the existing DataFrame
            if not df_existing[(df_existing["Subject"] == file) & (df_existing["Path"] == subdir)].empty:
                continue

            # Extract subject, sender and recipient information from the .msg file, with error control
            try:
                with extract_msg.Message(file_path) as msg:
                    subject = msg.subject
                    sender = msg.sender
                    recipients = msg.to
            except InvalidFileFormatError:
                    print(f"InvalidFileFormatError: Skipping file {file_path}")
                    continue

            # Remove leading number sequence and dash from file name, not necessary anymore if I use the subject
            file_name = re.sub(r"^\d+\s*-\s*", "", subject)

            # Remove "re" or "rv" prefix and any leading/trailing whitespace or ".msg" suffix from email subject
            subject = re.sub(r"^\s*(re|rv|fw[d]*)\s*:?\s*|\s*\.msg$", "", file_name, flags=re.IGNORECASE)

            # Add the modified subject, path, sender, recipients, and other information to the list
            file_list.append((subject.strip(), subdir, sender, recipients, None, pd.Timestamp.now()))

            # Increment the counter
            processed_emails += 1

            # Prints after each 1,000 emails processed
            if processed_emails % 1000 == 0:
                print (f"Processed {processed_emails} emails.")

            # Pause and ask for an "Enter" keypress after 10,000 emails processed
            if processed_emails % 10000 == 0:
                input(f"Processed {processed_emails} emails. Press 'Enter' to continue...")

# Create a DataFrame from the file list
df_new = pd.DataFrame(file_list, columns=["Subject", "Path", "Sender New", "Recipients New", "Archive New", "Date New"])

# Rename the "Archive", "Date Added", "Sender" and "Recipients" columns in the existing DataFrame to "Archive Existing", "Date Added Existing" "Sender Existing" and "Recipients Existing", respectively
df_existing.rename(columns={"Archive": "Archive Existing", "Date": "Date Existing", "Sender": "Sender Existing", "Recipients": "Recipients Existing"}, inplace=True)

# Merge the existing and new DataFrames on the "Email Subject" and "Path" columns, keeping all rows
df_all = pd.merge(df_existing, df_new, on=["Subject", "Path"], how="outer")

# Combine the "Sender Existing" and "Sender New" columns into a single "Sender" column
df_all["Sender"] = df_all["Sender Existing"].fillna(df_all["Sender New"])
df_all.drop(["Sender Existing", "Sender New"], axis=1, inplace=True)

# Combine the "Recipients Existing" and "Recipients New" columns into a single "Recipients" column
df_all["Recipients"] = df_all["Recipients Existing"].fillna(df_all["Recipients New"])
df_all.drop(["Recipients Existing", "Recipients New"], axis=1, inplace=True)

# Combine the "Archive Existing" and "Archive New" columns into a single "Archive" column
df_all["Archive"] = df_all["Archive Existing"].fillna(df_all["Archive New"])
df_all.drop(["Archive Existing", "Archive New"], axis=1, inplace=True)

# Combine the "Date Existing" and "Date New" columns into a single "Date" column
df_all["Date"] = df_all["Date Existing"].fillna(df_all["Date New"])
df_all.drop(["Date Existing", "Date New"], axis=1, inplace=True)

# Remove duplicate email subjects and paths
df_all.drop_duplicates(subset=["Subject", "Path"], keep="first", inplace=True)

# Rename the previous Excel file with an "-old" suffix
old_excel_path = os.path.splitext(excel_path)[0] + '-old' + os.path.splitext(excel_path)[1]

# If the "old" Excel file already exists, delete it
if os.path.exists(old_excel_path):
    os.remove(old_excel_path)

# If the previous Excel file exists, rename it to the "old" Excel file
if os.path.exists(excel_path):
    os.rename(excel_path, old_excel_path)

# Export the updated database to an Excel file
df_all.to_excel(excel_path, index=False)

# Apply the column widths to the final Excel file
if column_widths:
    wb_final = openpyxl.load_workbook(excel_path)
    ws_final = wb_final.active
    for i, width in enumerate(column_widths):
        ws_final.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width
    wb_final.save(excel_path)