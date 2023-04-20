# requirements: public
import os
import openpyxl
import pandas as pd
import re
import string
import win32com.client
from datetime import datetime

# requirements: custom functions
from utils import read_params_from_txt_file
from utils import get_first_explorer_folder_path

# Function to get the selected email from Outlook
def get_selected_email():
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    explorer = outlook.Application.ActiveExplorer()
    selection = explorer.Selection

    if len(selection) == 0:
        print("No email is selected.")
        return None

    return selection.Item(1)

# Function to get the next correlative number in the folder
def get_next_correlative_number(folder_path):
    files = os.listdir(folder_path)
    correlative_numbers = [int(re.findall(r'\d+', f)[0]) for f in files if re.findall(r'\d+', f)]

    if correlative_numbers:
        next_number = max(correlative_numbers) + 1
    else:
        next_number = 1

    return f"{next_number:03d}"

# Function to load the excel file
def load_excel_file(excel_path):
    return pd.read_excel(excel_path)

# Function to remove "re" or "rv" prefix and any leading/trailing whitespace or ".msg" suffix from email subject, to save in an excel
def sanitize_subject_re_rv(subject):
    subject = re.sub(r"^\s*(re|rv|fw[d]*)\s*:?\s*|\s*\.msg$", "", subject, flags=re.IGNORECASE)
    return subject

# Function to update the excel file with the recent
def update_excel_file(df, email_subject, file_path, sender, recipients, timestamp, excel_path):
    new_row = pd.DataFrame({"Subject": [email_subject], "Path": [file_path], "Sender": [sender], "Recipients": [recipients], "Date": [timestamp]})
    df = pd.concat([df, new_row], ignore_index=True)
    df = df.sort_values(by="Date", ascending=False)
    df = df.drop_duplicates(subset=["Path"], keep="first")
    df = df.head(10)
    df.to_excel(excel_path, index=False)

# Function to sanitize the email subject for saving as a msg file
def sanitize_subject(subject):
    valid_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
    sanitized_subject = ''.join(c for c in subject if c in valid_chars)
    return sanitized_subject.strip()

# Function to save email attachments
def save_attachments(email, folder_path, correlative_number):
    for attachment in email.Attachments:
        try:
            content_id = attachment.PropertyAccessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001E")
        except Exception:
            content_id = None

        # Check if the attachment has a ContentId, which is common for embedded images
        if content_id:
            continue

        file_path = os.path.join(folder_path, f"{correlative_number} - {attachment.FileName}")
        attachment.SaveAsFile(file_path)

# Function to save the email as a .msg file
def save_email_as_msg(email, folder_path, correlative_number):
    sanitized_subject = sanitize_subject(email.Subject)
    file_name = f"{correlative_number} - {sanitized_subject}.msg"
    file_path = os.path.join(folder_path, file_name)
    email.SaveAs(file_path)

# Function to archive the email
def archive_email(email):
    # Get the Gmail account
    outlook = email.Application.GetNamespace("MAPI")
    accounts = outlook.Accounts
    gmail_account = None

    for account in accounts:
        if "gmail" in account.SmtpAddress.lower():
            gmail_account = account
            break

    if not gmail_account:
        print("No Gmail account found.")
        return

    # Get the "All Mail" folder, which is the archive folder for Gmail
    all_mail_folder = gmail_account.DeliveryStore.GetRootFolder().Folders["[Gmail]"].Folders["Archivo"]

    if not all_mail_folder:
        print("No Gmail 'All Mail' folder found.")
        return

    # Move the email to the Gmail "All Mail" folder
    email.Move(all_mail_folder)

# Main execution

# Load the parameters from the text file
params_file_path = r"C:\Mis Datos en Local\temporal\python\email-automation-archive-params.txt"
params = read_params_from_txt_file(params_file_path)

# Get the active Windows Explorer instance's folder path
folder_path = get_first_explorer_folder_path()
if folder_path is None:
    print("No Windows Explorer instance found.")

# Load the email and archive
email = get_selected_email()
if email is None:
    print("No email found.")

if folder_path and email:

    # Calculate the next correlative number
    correlative_number = get_next_correlative_number(folder_path)
    print('correlative number = ' , correlative_number)

    # Save attachments
    save_attachments(email, folder_path, correlative_number)

    # Save email as .msg file
    save_email_as_msg(email, folder_path, correlative_number)

    # Archive the email (optional) uncomment the following line
    # archive_email(email)

    # Load excel_recent parameter
    excel_recent_path = params["recent_path"]

    # Open the excel file with pandas
    df = load_excel_file(excel_recent_path)

    # Get the column widths from the existing Excel file
    column_widths = []
    if os.path.exists(excel_recent_path):
        wb_existing = openpyxl.load_workbook(excel_recent_path)
        ws_existing = wb_existing.active
        column_widths = [ws_existing.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width for i in
                         range(len(df.columns))]

    # Add a new row to the excel file
    sanitized_subject = sanitize_subject_re_rv(email.subject)
    email_sender = email.SenderEmailAddress
    email_recipients = "; ".join([recipient.Address for recipient in email.Recipients])
    email_timestamp = datetime.now()

    update_excel_file(df, sanitized_subject, folder_path, email_sender, email_recipients, email_timestamp,
                      excel_recent_path)

    # Apply the column widths to the final Excel file
    if column_widths:
        wb_final = openpyxl.load_workbook(excel_recent_path)
        ws_final = wb_final.active
        for i, width in enumerate(column_widths):
            ws_final.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width
        wb_final.save(excel_recent_path)

else:
    print('No folder or email selected. Ending')