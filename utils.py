# requirements: public
import openpyxl
import win32com.client
import win32gui
from urllib.parse import unquote, urlsplit
import os
import pandas as pd
import pickle
from pathlib import Path
from datetime import datetime

# Custom function to replace special characters
def replace_special_chars(path):
    special_char_mapping = {
        '%E1': 'á',
        '%E9': 'é',
        '%ED': 'í',
        '%F3': 'ó',
        '%FA': 'ú',
        '%C1': 'Á',
        '%C9': 'É',
        '%CD': 'Í',
        '%D3': 'Ó',
        '%DA': 'Ú',
        '%F1': 'ñ',
        '%D1': 'Ñ',
        '%20': ' '
    }

    for key, value in special_char_mapping.items():
        path = path.replace(key, value)

    return path

# Function to get the path of the foreground windows explorer
def get_first_explorer_hwnd():
    # Get a list of all open windows
    windows = []
    win32gui.EnumWindows(lambda hwnd, windows: windows.append(hwnd), windows)

    # Find the first window that has a path to a directory in its title, taking into account network drives too
    for hwnd in windows:
        window_text = win32gui.GetWindowText(hwnd)
        if (":\\" in window_text or window_text.startswith("\\\\S555")) and not window_text.endswith(
                ".exe") and not window_text.endswith(".py"):
            return hwnd

    return None

def get_explorer_path_from_hwnd(target_hwnd):
    # Get all instances of Shell Windows
    shell_windows = win32com.client.Dispatch("Shell.Application").Windows()

    # Filter for Windows Explorer instances
    explorer_windows = [w for w in shell_windows if w.LocationURL.startswith("file:///")]

    if not explorer_windows:
        print("No Windows Explorer instances found.")
        return None

    # Iterate through explorer_windows to find the matching HWND and return the folder path
    for window in explorer_windows:
        hwnd = window.HWND
        if hwnd == target_hwnd:
            print(window.LocationURL)
            url_parts = urlsplit(window.LocationURL)
            folder_path = url_parts.path
            folder_path = folder_path[1:] if folder_path.startswith('/') else folder_path
            folder_path = folder_path.replace('/', '\\')
            folder_path = replace_special_chars(folder_path)
            return folder_path

    print("No matching Windows Explorer instance found.")
    return None

def get_first_explorer_folder_path():
    # Get the HWND of the first Windows Explorer instance with a path in its title
    first_explorer_hwnd = get_first_explorer_hwnd()

    # If no matching HWND is found, print an error message and return None
    if first_explorer_hwnd is None:
        print("No Windows Explorer instance found with a path in its title.")
        return None

    # Get the folder path of the Windows Explorer instance with the matching HWND
    folder_path = get_explorer_path_from_hwnd(first_explorer_hwnd)

    # If a folder path is found, print the HWND and folder path, then return the folder path as a string
    if folder_path:
        print(f"Window handle: {first_explorer_hwnd}, Folder path: {folder_path}")
        return folder_path

    return None

# Function to read the parameters from the txt file
def read_params_from_txt_file(file_path):
    params = {}
    with open(file_path, 'r') as f:
        for line in f:
            if line.strip():
                key, value = line.strip().split(" = ", 1)
                params[key.strip()] = value.strip()
    return params

# Function to open an excel file or a pickle, if found. If not found, creates the pickle
def read_excel_or_pickle(excel_file_path, pickle_file_path, sheet_name=None, usecols=None, engine=None):
    excel_file = Path(excel_file_path)
    pickle_file = Path(pickle_file_path)

    start_time = datetime.now()
    print(f"Starting data loading process at {start_time}")
    if pickle_file.exists() and pickle_file.stat().st_mtime > excel_file.stat().st_mtime:
        print(f"Loading data from pickle file {pickle_file}")
        with open(pickle_file, 'rb') as f:
            df = pickle.load(f)
    else:
        print(f"Loading data from Excel file and creating a pickle file {excel_file}")
        df = pd.read_excel(excel_file, sheet_name=sheet_name, usecols=usecols, engine=engine)
        with open(pickle_file, 'wb') as f:
            pickle.dump(df, f)

    end_time = datetime.now()
    print(f"Data loading process finished at {end_time}")

    duration = end_time - start_time
    print(f"Total duration of the data loading: {duration}")

    # If no sheet_name is specified and df is a dictionary, return the first DataFrame
    if sheet_name is None and isinstance(df, dict):
        first_sheet = list(df.values())[0]
        return first_sheet

    return df

# Get the column widths from the existing Excel file, initializing column_widths as an empty list first
def get_column_widths(excel_path):
    if os.path.exists(excel_path):
        wb_existing = openpyxl.load_workbook(excel_path)
        ws_existing = wb_existing.active

        # Find the last column with content
        last_col = ws_existing.max_column

        # Get column widths for columns with content
        column_widths = [ws_existing.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width for i in range(last_col)]
    return column_widths

# Apply column widths to an excel file (requires the column_widths)
def apply_column_widths(excel_path, column_widths):
    if column_widths:
        wb_final = openpyxl.load_workbook(excel_path)
        ws_final = wb_final.active
        for i, width in enumerate(column_widths):
            ws_final.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width
        wb_final.save(excel_path)