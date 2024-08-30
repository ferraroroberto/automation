import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk, filedialog, messagebox

# Hardcoded column names
hierarchy1_code = 'COD_DT'
hierarchy1_name = 'DES_DT'
hierarchy2_code = 'COD_DG'
hierarchy2_name = 'DES_DG'
hierarchy3_code = 'COD_DAN'
hierarchy3_name = 'DES_DAN'
hierarchy4_code = 'PK_CENTRO'
hierarchy4_name = 'DES_CENTRO_GES'
type_name = 'FK_TIPO_CENTRO_GES'

# List of required columns
required_columns = [
    hierarchy1_code, hierarchy1_name,
    hierarchy2_code, hierarchy2_name,
    hierarchy3_code, hierarchy3_name,
    hierarchy4_code, hierarchy4_name,
    type_name
]

def select_excel_file():
    """Open a file dialog to select an Excel file."""
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    if not file_path:
        messagebox.showerror("Error", "No file selected.")
        raise FileNotFoundError("No file selected.")
    return file_path

def get_user_input(default_values):
    """Prompt the user to manually input sheet and column names or use default values."""
    user_input = messagebox.askyesno("Input Required", "Do you want to manually input sheet and column names?")
    if user_input:
        sheet_names = {}
        for key in default_values.keys():
            value = input(f"Enter the name for {key} (default: {default_values[key]}): ")
            sheet_names[key] = value if value else default_values[key]
        return sheet_names
    else:
        return default_values

def transform_hierarchy(input_file, sheet_names):
    """Transform the hierarchy data from the Excel file."""
    wb = load_workbook(input_file)
    sheet = wb[wb.sheetnames[0]]  # Get the first sheet
    data = pd.DataFrame(sheet.values)
    data.columns = data.iloc[0]
    data = data[1:]

    # Ensure required columns are present
    for col in required_columns:
        if col not in data.columns:
            raise ValueError(f"Missing required column: {col}")

    # Sort the data by the hierarchy columns
    data = data.sort_values(by=[hierarchy1_code, hierarchy2_code, hierarchy3_code, hierarchy4_code])

    # Transform the hierarchy
    transformed_data = []
    seen_codes = set()

    for _, row in data.iterrows():
        # Handle hierarchy1 (DT)
        if row[hierarchy1_code] not in seen_codes:
            transformed_data.append({
                'hierarchy1_code': row[hierarchy1_code],
                'hierarchy2_code': row[hierarchy1_code],
                'hierarchy3_code': row[hierarchy1_code],
                'hierarchy4_code': row[hierarchy1_code],
                'center_code': row[hierarchy1_code],
                'center_name': row[hierarchy1_name],
                'center_type': 'DT'
            })
            seen_codes.add(row[hierarchy1_code])
        
        # Handle hierarchy2 (DC)
        if pd.notna(row[hierarchy2_code]) and row[hierarchy2_code] not in seen_codes:
            transformed_data.append({
                'hierarchy1_code': row[hierarchy1_code],
                'hierarchy2_code': row[hierarchy2_code],
                'hierarchy3_code': row[hierarchy2_code],
                'hierarchy4_code': row[hierarchy2_code],
                'center_code': row[hierarchy2_code],
                'center_name': row[hierarchy2_name],
                'center_type': 'DC'
            })
            seen_codes.add(row[hierarchy2_code])
        
        # Handle hierarchy3 (DAN)
        if pd.notna(row[hierarchy3_code]) and row[hierarchy3_code] not in seen_codes:
            transformed_data.append({
                'hierarchy1_code': row[hierarchy1_code],
                'hierarchy2_code': row[hierarchy2_code],
                'hierarchy3_code': row[hierarchy3_code],
                'hierarchy4_code': row[hierarchy3_code],
                'center_code': row[hierarchy3_code],
                'center_name': row[hierarchy3_name],
                'center_type': 'DAN'
            })
            seen_codes.add(row[hierarchy3_code])
        
        # Handle hierarchy4 (OFI)
        if pd.notna(row[hierarchy4_code]) and row[hierarchy4_code] not in seen_codes:
            transformed_data.append({
                'hierarchy1_code': row[hierarchy1_code],
                'hierarchy2_code': row[hierarchy2_code],
                'hierarchy3_code': row[hierarchy3_code],
                'hierarchy4_code': row[hierarchy4_code],
                'center_code': row[hierarchy4_code],
                'center_name': row[hierarchy4_name],
                'center_type': 'OFI'
            })
            seen_codes.add(row[hierarchy4_code])

    # Remove duplicates
    transformed_df = pd.DataFrame(transformed_data).drop_duplicates()

    # Save the transformed data to a new Excel file
    output_file = input_file.replace('.xlsx', '_transformed.xlsx')
    transformed_df.to_excel(output_file, index=False)
    print(f"Transformed data saved to {output_file}")

# Example usage
if __name__ == "__main__":
    default_values = {
        'hierarchy1_code': hierarchy1_code,
        'hierarchy1_name': hierarchy1_name,
        'hierarchy2_code': hierarchy2_code,
        'hierarchy2_name': hierarchy2_name,
        'hierarchy3_code': hierarchy3_code,
        'hierarchy3_name': hierarchy3_name,
        'hierarchy4_code': hierarchy4_code,
        'hierarchy4_name': hierarchy4_name,
        'type_name': type_name
    }

    try:
        input_file = select_excel_file()
        sheet_names = get_user_input(default_values)
        transform_hierarchy(input_file, sheet_names)
    except FileNotFoundError as e:
        print(e)
    except ValueError as e:
        print(e)
